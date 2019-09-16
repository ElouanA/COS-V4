from projet import Projet
from canevas import Canevas
from jure import Jure
from etudiant import Etudiant
from openpyxl import *
from critere import Critere
from appreciation import Appreciation
import string
from openpyxl.styles import colors
from openpyxl.styles import Font, Color, PatternFill, Border, Alignment


## I - Outils de gestion des lignes et colonnes
def index (row, column):
    List=list(string.ascii_uppercase)
    return (List[column-1] + str(row))

def indextoletter(column):
    List=list(string.ascii_uppercase)
    res=''
    ecriture=[]
    j=1
    while (column//(26**j)!=0): 
        j+=1
    while j!=0:
        ecriture.append(column//(26**(j-1)))
        column=column%(26**(j-1))
        j-=1
    for i in range (0,len(ecriture)):
        res+=List[ecriture[i]-1]
    return(res)

## II - Définition de la classe Cos


class Cos:
    """ Classe définissant la façade du programme """
    
    def __init__(self,wb):
        self.canevas = None
        self.jures = []
        self.projets = []
        self.wb = wb
        self.wbout = Workbook()
        
    def add_canevas(self,canevas):
        if isinstance(canevas, Canevas):
            self.canevas.append(canevas)
        
    def add_projet(self,projet):
        if isinstance(projet, Projet):
            if (self.chercher_projet_by_name(projet.nom)==None):
                self.projets.append(projet)
        
    def add_jure(self,jure):
        if isinstance(jure, Jure):
            self.jures.append(jure)
            
    def chercher_projet_by_name(self,name):
        projet = None
        for i in self.projets :
            if (i.nom == name):
                projet = i
        return projet
        
    def chercher_projet_by_id(self,id):
        projet = None
        for i in self.projets :
            if (i.id == id):
                projet = i
        return projet
    
    def to_string(self):
        res = self.canevas.to_string()
        for i in self.jures:
            res += i.to_string()
            res += "\n"
        for j in self.projets:
            res += j.to_string()
            res += "\n"
            
        return(res)

## III - Récupération des données de la feuille entrée par jury


    def collect_etudiants_et_projet_from_jury(self):
        sheet=self.wb[self.wb.sheetnames[2]]
        i=4
        while (sheet.cell(row=i, column=1).value != None) : #tant qu'il y a des élèves dans la liste
            if self.projets.count( self.chercher_projet_by_name(sheet.cell(row=i,column = 3).value)) == 0:
                self.add_jure(Etudiant(sheet.cell(row=i, column=1).value,sheet.cell(row=i, column=2).value,len(self.projets)+1))
                self.add_projet(Projet(sheet.cell(row=i,column = 3).value,len(self.projets)+1))
               
            else:
                self.add_jure(Etudiant(sheet.cell(row=i, column=1).value,sheet.cell(row=i, column=2).value,self.chercher_projet_by_name(sheet.cell(row=i,column = 3).value).id))
                
            i+=1
    
    def collect_jures_from_jury(self):
        sheet=self.wb[self.wb.sheetnames[3]]
        i=4
        while (sheet.cell(row=i, column=1).value != None) : #tant qu'il y a des jurés dans la liste
            self.add_jure(Jure(sheet.cell(row=i, column=1).value,sheet.cell(row=i, column=2).value))
            i+=1
    
    def getCritere(self,row): 
        sheet=self.wb[self.wb.sheetnames[0]]
        titre = sheet.cell(row = row , column=1).value #titre
        description = sheet.cell(row=row +1, column= 2).value #description
        points = sheet.cell(row=row , column = 3).value #nbpoints
        if description == None:
            hauteur = None
        else: 
            hauteur = sheet.row_dimensions[row+1].height
        return(Critere(titre,description,points,hauteur))
        
    def getAppreciation(self,row):
        sheet=self.wb[self.wb.sheetnames[0]]
        titre = sheet.cell(row = row , column=1).value #titre
        nbPoints = sheet.cell(row = row , column=3).value #nbPoints
        partDeLaNote = sheet.cell(row = row , column=4).value #partDeLaNote
        i=row+1
        appreciation = Appreciation(titre, nbPoints, partDeLaNote)
        while (sheet.cell(row=i,column = 1).value != None): 
            
            appreciation.add_critere(self.getCritere(i)) 
            if (self.getCritere(i).description == None): 
                i+=1
            else: 
                i+=2
        return(appreciation) 
        
    def list_appreciations_from_jury(self):
        appreciations = []
        sheet=self.wb[self.wb.sheetnames[0]]
        i=4
        while (sheet.cell(row=i,column = 1).value != None): 
            temp = self.getAppreciation(i)
            appreciations.append(temp)
            for j in range (0, len(temp.criteres)):
                if (temp.criteres[j].description == None):
                    i+=1
                else: 
                    i+=2
            i+=2
        return(appreciations)
            
    def define_canevas(self):
        sheet=self.wb[self.wb.sheetnames[0]]
        appreciations = self.list_appreciations_from_jury()
        try:
            widthCritere = int(sheet.column_dimensions['B'].width)
        except TypeError:
            widthCritere = 20
        try:
            widthMarge1 = int(sheet.column_dimensions['A'].width)
        except TypeError:
            widthMarge1 = 20 
        try:
            widthPoints = int(sheet.column_dimensions['C'].width)
        except TypeError:
            widthPoints = 20
        try:
            widthPourcentage = int(sheet.column_dimensions['D'].width)
        except TypeError:
            widthPourcentage = 20
        try:
            widthNoteFinale =  int(sheet.column_dimensions['F'].width)
        except TypeError:
            widthNoteFinale = 20
        try:
            widthMarge2 =  int(sheet.column_dimensions['E'].width)
        except TypeError:
            widthMarge2 = 20
        try:
            widthMarge3 =  int(sheet.column_dimensions['G'].width)
        except TypeError:
            widthMarge3 = 20 
        try:
            widthCommentaireJure =  int(sheet.column_dimensions['I'].width)
        except TypeError:
            widthCommentaireJure = 20
        try:
            widthNoteJure = int(sheet.column_dimensions['H'].width)
        except TypeError:
            widthNoteJure=20
        colors = [sheet.cell(row=4, column=1).fill,sheet.cell(row=4, column=4).fill,sheet.cell(row=4, column=6).fill,sheet.cell(row=2, column=8).fill,sheet.cell(row=2, column=10).fill]
        try:
            heightJure = int(sheet.row_dimensions[2].height)
        except TypeError:
            heightJure= 20 
        fontTitre = sheet.cell(row= 1 ,column=1).font
        fontAppreciation = sheet.cell(row= 4 ,column=1).font
        fontTitreCritere = sheet.cell(row= 5 ,column=1).font
        fontDescriptionCritere = sheet.cell(row= 6 ,column=2).font
        fontNoteCritere = sheet.cell(row= 5 ,column=3).font
        fontNoteAppreciation = sheet.cell(row= 4 ,column=3).font
        fontNotePourcentage = sheet.cell(row= 4 ,column= 4).font
        fontPourcentage = sheet.cell(row= 2 ,column=4).font
        fontNoteFinale = sheet.cell(row= 4 ,column=6).font
        fontTitreNoteFinale = sheet.cell(row= 3 ,column=6).font
        fontTitreJure = sheet.cell(row= 2 ,column=8).font
        fontColonneCommentaireJure = sheet.cell(row= 3 ,column=9).font
        fontCommentaireJure = sheet.cell(row= 4 ,column=9).font
        fontColonneNoteJure = sheet.cell(row= 3 ,column=8).font
        fontNoteAppreciationJure = sheet.cell(row= 4 ,column=8).font
        fontNoteCritereJure = sheet.cell(row= 5 ,column=8).font
        titreColonneCommentaireJure = sheet.cell(row=3, column=9).value
        titreColonneNoteJure = sheet.cell(row=3, column=8).value
        titrePourcentage = sheet.cell(row=2, column=4).value
        titreNoteFinale = sheet.cell(row=3, column=6).value
        sheet=self.wb[self.wb.sheetnames[2]]
        try:
            width_nom = int(sheet.column_dimensions['A'].width)
        except TypeError:
            width_nom = 20
        try:
            width_prenom = int(sheet.column_dimensions['B'].width)
        except TypeError:
            width_prenom = 20
        try: 
            width_projet = int(sheet.column_dimensions['C'].width)
        except: 
            width_projet = 20
        border = sheet.cell(row=3,column=1).border
        
        
        self.canevas = Canevas(widthCritere,widthMarge1,widthPoints,widthPourcentage,widthNoteFinale,widthMarge2,widthMarge3,widthCommentaireJure,widthNoteJure,colors, heightJure,fontTitre,fontAppreciation,fontTitreCritere,fontDescriptionCritere,fontNoteCritere,fontNoteAppreciation,fontNotePourcentage,fontPourcentage,fontNoteFinale,fontTitreNoteFinale,fontTitreJure,fontColonneCommentaireJure,fontCommentaireJure,fontColonneNoteJure,fontNoteAppreciationJure,fontNoteCritereJure,titreColonneCommentaireJure,titreColonneNoteJure,titrePourcentage,titreNoteFinale,width_nom,width_prenom, width_projet,border)
        
        self.canevas.appreciations = self.list_appreciations_from_jury()

## IV - Ajout des informations au cos

        
    def initialize_cos_jury(self):
        self.collect_jures_from_jury()
        self.collect_etudiants_et_projet_from_jury()
        self.define_canevas()
        
    
## V - Génération de la feuille sortie par Jury
        
    def generate_synthese_etudiant(self):
        ws = self.wbout.create_sheet("Synthese_etudiants")
        ws.merge_cells(start_row = 1,start_column = 1, end_row = 1, end_column = 4)
        ws.column_dimensions['A'].width = self.canevas.width_nom
        ws.column_dimensions['B'].width = self.canevas.width_prenom
        ws.column_dimensions['C'].width = self.canevas.width_projet
        ws.column_dimensions['D'].width = 38
        ws.column_dimensions['E'].width = 24
        ws.column_dimensions['F'].width = 40
        ws.cell(row =1, column = 1).value =  "Synthese notes étudiants"
        ws.cell(row =1, column = 1).alignment = Alignment(wrap_text=True)
        ws.row_dimensions[3].height = 40
        ws.cell(row =3, column = 1).value = "NOM"
        ws.cell(row =3, column = 1).alignment = Alignment(wrap_text=True)
        ws.cell(row =3, column = 2).value = "Prenom"
        ws.cell(row =3, column = 2).alignment = Alignment(wrap_text=True)
        ws.cell(row =3, column = 3).value = "Projet"
        ws.cell(row =3, column = 3).alignment = Alignment(wrap_text=True)
        j=4 
        for appreciation in self.canevas.appreciations: 
            ws.cell(row =3, column = j).value = appreciation.titre
            ws.cell(row =3, column = j).alignment = Alignment(wrap_text=True)
            i=4
            for etudiant in self.jures:
                if isinstance(etudiant,Etudiant):
                    ws.cell(row=i,column=j).value="Projet_"+self.chercher_projet_by_id(etudiant.idprojet).nom.replace(" ","")+"!"
                    i+=1
            j+=1
        finappreciation=j
        ws.cell(row =3, column = finappreciation).value = 'Bonus évaluation par les pairs (NB : colonne à mettre seulement si paramètre Generer_aussi_pour_etudiants vaut "Oui")'
        ws.cell(row =3, column = finappreciation).alignment = Alignment(wrap_text=True)
        for i in range(finappreciation) :
            ws.cell(row = 3, column = i+1).border = Border(self.canevas.border.left, self.canevas.border.right, self.canevas.border.top, self.canevas.border.bottom)
        
        j=4
        for etudiant in self.jures:
            if isinstance(etudiant, Etudiant):
                projet = self.chercher_projet_by_id(etudiant.idprojet)
                ws.cell(row=j,column=1).value= etudiant.nom
                ws.cell(row =j, column = 1).alignment = Alignment(wrap_text=True)
                ws.cell(row=j,column=2).value= etudiant.prenom
                ws.cell(row =j, column = 2).alignment = Alignment(wrap_text=True)
                ws.cell(row=j,column=3).value= projet.nom
                ws.cell(row =j, column = 3).alignment = Alignment(wrap_text=True)
                for i in range(finappreciation) :
                    ws.cell(row = j, column = i+1).border = Border(self.canevas.border.left, self.canevas.border.right, self.canevas.border.top, self.canevas.border.bottom)
                j+=1
                
        
        self.wbout.save("FichierSortie_parJury.xlsx")
        
    
        
    def generate_projets_sheets(self):
        
        nb_lignes = 2*len(self.canevas.appreciations)-1
        for appreciation in self.canevas.appreciations:
            for critere in appreciation.criteres:
                if critere.description == None:
                    nb_lignes += 1
                else:
                    nb_lignes += 2
                    
                    
        nb_colonnes = 0
        for jure in self.jures:
            if not isinstance(jure, Etudiant):
                nb_colonnes+=1
        nb_colonnes = nb_colonnes*2
        
        
        first_row = 2
        first_column = 8
        wssyntheseetudiant=self.wbout[self.wbout.sheetnames[0]]
        wssyntheseprojet=self.wbout[self.wbout.sheetnames[1]]
        numeroprojet=0
        for projet in self.projets:
            ws = self.wbout.create_sheet("Projet_"+projet.nom.replace(" ",""))
            #for ligne in range(nb_lignes):
                #for colonne in range(nb_colonnes):
                    #ws.cell(row = first_row+ligne, column = first_column+colonne).border = Border(self.canevas.border.left, self.canevas.border.right, self.canevas.border.top, self.canevas.border.bottom)
            ws.merge_cells(start_row = 1,start_column = 1, end_row = 1, end_column = 4)
            ws.row_dimensions[2].height=self.canevas.heightJure
            ws.column_dimensions['A'].width=self.canevas.widthMarge1
            ws.column_dimensions['B'].width=self.canevas.widthCritere
            ws.column_dimensions['C'].width=self.canevas.widthPoints
            ws.column_dimensions['D'].width=self.canevas.widthPourcentage
            ws.column_dimensions['E'].width=self.canevas.widthMarge2
            ws.column_dimensions['F'].width=self.canevas.widthNoteFinale
            ws.column_dimensions['G'].width=self.canevas.widthMarge3
            ws.cell(row=1,column=1).value= "Projet_"+projet.nom
            ws.cell(row =1, column = 1).alignment = Alignment(wrap_text=True)
            ws.cell(row=1,column=1).font=Font(name=self.canevas.fontTitre.name, charset=self.canevas.fontTitre.charset, family=self.canevas.fontTitre.family, b=self.canevas.fontTitre.b, i=self.canevas.fontTitre.i, strike=self.canevas.fontTitre.strike, outline=self.canevas.fontTitre.outline, shadow=self.canevas.fontTitre.shadow, condense=self.canevas.fontTitre.condense, color=self.canevas.fontTitre.color, extend=self.canevas.fontTitre.extend, sz=self.canevas.fontTitre.size, u=self.canevas.fontTitre.u, vertAlign=self.canevas.fontTitre.vertAlign, scheme=self.canevas.fontTitre.scheme)
            ws.cell(row=2,column = 4).value = self.canevas.titrePourcentage
            #ws.cell(row = 2, column = 4).border = Border(self.canevas.border.left, self.canevas.border.right, self.canevas.border.top, self.canevas.border.bottom)
            ws.cell(row =2, column = 4).alignment = Alignment(wrap_text=True)
            ws.cell(row=2,column = 4).fill = PatternFill(patternType=self.canevas.colors[1].patternType, fgColor=self.canevas.colors[1].fgColor, bgColor=self.canevas.colors[1].bgColor) 
            ws.cell(row=2,column=4).font=Font(name=self.canevas.fontPourcentage.name, charset=self.canevas.fontPourcentage.charset, family=self.canevas.fontPourcentage.family, b=self.canevas.fontPourcentage.b, i=self.canevas.fontPourcentage.i, strike=self.canevas.fontPourcentage.strike, outline=self.canevas.fontPourcentage.outline, shadow=self.canevas.fontPourcentage.shadow, condense=self.canevas.fontPourcentage.condense, color=self.canevas.fontPourcentage.color, extend=self.canevas.fontPourcentage.extend, sz=self.canevas.fontPourcentage.size, u=self.canevas.fontPourcentage.u, vertAlign=self.canevas.fontPourcentage.vertAlign, scheme=self.canevas.fontPourcentage.scheme)
            ws.cell(row=3,column = 6).value = self.canevas.titreNoteFinale
            #ws.cell(row = 3, column = 6).border = Border(self.canevas.border.left, self.canevas.border.right, self.canevas.border.top, self.canevas.border.bottom)
            ws.cell(row =3, column = 6).alignment = Alignment(wrap_text=True)
            ws.cell(row=3,column=6).font=Font(name=self.canevas.fontTitreNoteFinale.name, charset=self.canevas.fontTitreNoteFinale.charset, family=self.canevas.fontTitreNoteFinale.family, b=self.canevas.fontTitreNoteFinale.b, i=self.canevas.fontTitreNoteFinale.i, strike=self.canevas.fontTitreNoteFinale.strike, outline=self.canevas.fontTitreNoteFinale.outline, shadow=self.canevas.fontTitreNoteFinale.shadow, condense=self.canevas.fontTitreNoteFinale.condense, color=self.canevas.fontTitreNoteFinale.color, extend=self.canevas.fontTitreNoteFinale.extend, sz=self.canevas.fontTitreNoteFinale.size, u=self.canevas.fontTitreNoteFinale.u, vertAlign=self.canevas.fontTitreNoteFinale.vertAlign, scheme=self.canevas.fontTitreNoteFinale.scheme)
            i=4
            indiceslignenote=[]
            numeroappreciation=0
            for appreciation in self.canevas.appreciations:
                debutappreciation=i
                indiceslignenote.append(debutappreciation)
                #ws.cell(row = debutappreciation, column = 6).border = Border(self.canevas.border.left, self.canevas.border.right, self.canevas.border.top, self.canevas.border.bottom)
                ws.cell(row =debutappreciation, column = 6).alignment = Alignment(wrap_text=True)
                ws.cell(row=debutappreciation,column = 6).fill = PatternFill(patternType=self.canevas.colors[0].patternType, fgColor=self.canevas.colors[0].fgColor, bgColor=self.canevas.colors[0].bgColor)
                ws.cell(row=debutappreciation,column=6).font=Font(name=self.canevas.fontNoteFinale.name, charset=self.canevas.fontNoteFinale.charset, family=self.canevas.fontNoteFinale.family, b=self.canevas.fontNoteFinale.b, i=self.canevas.fontNoteFinale.i, strike=self.canevas.fontNoteFinale.strike, outline=self.canevas.fontNoteFinale.outline, shadow=self.canevas.fontNoteFinale.shadow, condense=self.canevas.fontNoteFinale.condense, color=self.canevas.fontNoteFinale.color, extend=self.canevas.fontNoteFinale.extend, sz=self.canevas.fontNoteFinale.size, u=self.canevas.fontNoteFinale.u, vertAlign=self.canevas.fontNoteFinale.vertAlign, scheme=self.canevas.fontNoteFinale.scheme)
                ws.merge_cells(start_row = i,start_column = 1, end_row = i, end_column = 2)
                ws.cell(row=i,column = 1).value = appreciation.titre
                #ws.cell(row = i, column = 1).border = Border(self.canevas.border.left, self.canevas.border.right, self.canevas.border.top, self.canevas.border.bottom)
                #ws.cell(row = i, column = 2).border = Border(self.canevas.border.left, self.canevas.border.right, self.canevas.border.top, self.canevas.border.bottom)
                ws.cell(row =i, column = 1).alignment = Alignment(wrap_text=True)
                ws.cell(row=i,column=1).font=Font(name=self.canevas.fontAppreciation.name, charset=self.canevas.fontAppreciation.charset, family=self.canevas.fontAppreciation.family, b=self.canevas.fontAppreciation.b, i=self.canevas.fontAppreciation.i, strike=self.canevas.fontAppreciation.strike, outline=self.canevas.fontAppreciation.outline, shadow=self.canevas.fontAppreciation.shadow, condense=self.canevas.fontAppreciation.condense, color=self.canevas.fontAppreciation.color, extend=self.canevas.fontAppreciation.extend, sz=self.canevas.fontAppreciation.size, u=self.canevas.fontAppreciation.u, vertAlign=self.canevas.fontAppreciation.vertAlign, scheme=self.canevas.fontAppreciation.scheme)
                ws.cell(row=i,column = 1).fill = PatternFill(patternType=self.canevas.colors[0].patternType, fgColor=self.canevas.colors[0].fgColor, bgColor=self.canevas.colors[0].bgColor)
                ws.cell(row=i,column = 3).value = appreciation.nbPoints
                #ws.cell(row = i, column = 3).border = Border(self.canevas.border.left, self.canevas.border.right, self.canevas.border.top, self.canevas.border.bottom)
                ws.cell(row =i, column = 3).alignment = Alignment(wrap_text=True)
                ws.cell(row=i,column=3).font=Font(name=self.canevas.fontNoteAppreciation.name, charset=self.canevas.fontNoteAppreciation.charset, family=self.canevas.fontNoteAppreciation.family, b=self.canevas.fontNoteAppreciation.b, i=self.canevas.fontNoteAppreciation.i, strike=self.canevas.fontNoteAppreciation.strike, outline=self.canevas.fontNoteAppreciation.outline, shadow=self.canevas.fontNoteAppreciation.shadow, condense=self.canevas.fontNoteAppreciation.condense, color=self.canevas.fontNoteAppreciation.color, extend=self.canevas.fontNoteAppreciation.extend, sz=self.canevas.fontNoteAppreciation.size, u=self.canevas.fontNoteAppreciation.u, vertAlign=self.canevas.fontNoteAppreciation.vertAlign, scheme=self.canevas.fontNoteAppreciation.scheme)
                ws.cell(row=i,column = 3).fill = PatternFill(patternType=self.canevas.colors[0].patternType, fgColor=self.canevas.colors[0].fgColor, bgColor=self.canevas.colors[0].bgColor) 
                ws.cell(row=i,column = 4).value = str(int(float(appreciation.partDeLaNote)*100)) + '%'
                #ws.cell(row = i, column = 4).border = Border(self.canevas.border.left, self.canevas.border.right, self.canevas.border.top, self.canevas.border.bottom)
                ws.cell(row =i, column = 4).alignment = Alignment(wrap_text=True) 
                ws.cell(row=i,column=4).font=Font(name=self.canevas.fontNotePourcentage.name, charset=self.canevas.fontNotePourcentage.charset, family=self.canevas.fontNotePourcentage.family, b=self.canevas.fontNotePourcentage.b, i=self.canevas.fontNotePourcentage.i, strike=self.canevas.fontNotePourcentage.strike, outline=self.canevas.fontNotePourcentage.outline, shadow=self.canevas.fontNotePourcentage.shadow, condense=self.canevas.fontNotePourcentage.condense, color=self.canevas.fontNotePourcentage.color, extend=self.canevas.fontNotePourcentage.extend, sz=self.canevas.fontNotePourcentage.size, u=self.canevas.fontNotePourcentage.u, vertAlign=self.canevas.fontNotePourcentage.vertAlign, scheme=self.canevas.fontNotePourcentage.scheme)
                ws.cell(row=i,column = 4).fill = PatternFill(patternType=self.canevas.colors[1].patternType, fgColor=self.canevas.colors[1].fgColor, bgColor=self.canevas.colors[1].bgColor) 
                i+=1
                for critere in appreciation.criteres:
                    ws.merge_cells(start_row = i,start_column = 1, end_row =i, end_column = 2)
                    ws.cell(row=i,column = 1).value = critere.titre
                    #ws.cell(row = i, column = 1).border = Border(self.canevas.border.left, self.canevas.border.right, self.canevas.border.top, self.canevas.border.bottom)
                    #ws.cell(row = i, column = 2).border = Border(self.canevas.border.left, self.canevas.border.right, self.canevas.border.top, self.canevas.border.bottom)
                    ws.cell(row =i, column = 1).alignment = Alignment(wrap_text=True)
                    ws.cell(row=i,column=1).font=Font(name=self.canevas.fontTitreCritere.name, charset=self.canevas.fontTitreCritere.charset, family=self.canevas.fontTitreCritere.family, b=self.canevas.fontTitreCritere.b, i=self.canevas.fontTitreCritere.i, strike=self.canevas.fontTitreCritere.strike, outline=self.canevas.fontTitreCritere.outline, shadow=self.canevas.fontTitreCritere.shadow, condense=self.canevas.fontTitreCritere.condense, color=self.canevas.fontTitreCritere.color, extend=self.canevas.fontTitreCritere.extend, sz=self.canevas.fontTitreCritere.size, u=self.canevas.fontTitreCritere.u, vertAlign=self.canevas.fontTitreCritere.vertAlign, scheme=self.canevas.fontTitreCritere.scheme)
                    ws.cell(row=i,column = 3).value = critere.points
                    #ws.cell(row = i, column = 3).border = Border(self.canevas.border.left, self.canevas.border.right, self.canevas.border.top, self.canevas.border.bottom)
                    ws.cell(row =i, column = 3).alignment = Alignment(wrap_text=True)
                    ws.cell(row=i,column=3).font=Font(name=self.canevas.fontNoteCritere.name, charset=self.canevas.fontNoteCritere.charset, family=self.canevas.fontNoteCritere.family, b=self.canevas.fontNoteCritere.b, i=self.canevas.fontNoteCritere.i, strike=self.canevas.fontNoteCritere.strike, outline=self.canevas.fontNoteCritere.outline, shadow=self.canevas.fontNoteCritere.shadow, condense=self.canevas.fontNoteCritere.condense, color=self.canevas.fontNoteCritere.color, extend=self.canevas.fontNoteCritere.extend, sz=self.canevas.fontNoteCritere.size, u=self.canevas.fontNoteCritere.u, vertAlign=self.canevas.fontNoteCritere.vertAlign, scheme=self.canevas.fontNoteCritere.scheme)
                    j=8
                    debutjure=j
                    for jure in self.jures:
                        if not isinstance(jure,Etudiant):
                            ws.cell(row=i,column=j).value="=Jury_"+jure.nom.replace(" ","")+'_'+jure.prenom.replace(" ","")+"!"+indextoletter(2*numeroprojet+8)+str(i)
                            j+=2
                    finjure=j
                    if (critere.description == None):
                        i+=1
                    else:
                        #ws.cell(row = i+1, column = 3).border = Border(self.canevas.border.left, self.canevas.border.right, self.canevas.border.top, self.canevas.border.bottom)
                        j=8 #indice de colonnes parcourues à travers les jurés
                        for jure in self.jures:
                            if not isinstance(jure, Etudiant) :
                                ws.merge_cells(start_row = i,start_column = j, end_row = i+1, end_column = j)
                                j+=2
                        ws.merge_cells(start_row = i,start_column = 3, end_row = i+1, end_column = 3)
                        ws.row_dimensions[i+1].height = critere.hauteur
                        ws.cell(row=i+1,column = 2).value = critere.description
                        #ws.cell(row = i+1, column = 2).border = Border(self.canevas.border.left, self.canevas.border.right, self.canevas.border.top, self.canevas.border.bottom)
                        ws.cell(row =i+1, column = 2).alignment = Alignment(wrap_text=True)
                        ws.cell(row=i+1,column=2).font=Font(name=self.canevas.fontDescriptionCritere.name, charset=self.canevas.fontDescriptionCritere.charset, family=self.canevas.fontDescriptionCritere.family, b=self.canevas.fontDescriptionCritere.b, i=self.canevas.fontDescriptionCritere.i, strike=self.canevas.fontDescriptionCritere.strike, outline=self.canevas.fontDescriptionCritere.outline, shadow=self.canevas.fontDescriptionCritere.shadow, condense=self.canevas.fontDescriptionCritere.condense, color=self.canevas.fontDescriptionCritere.color, extend=self.canevas.fontDescriptionCritere.extend, sz=self.canevas.fontDescriptionCritere.size, u=self.canevas.fontDescriptionCritere.u, vertAlign=self.canevas.fontDescriptionCritere.vertAlign, scheme=self.canevas.fontDescriptionCritere.scheme)
                        i+=2
                finappreciation=i-1
                ws.merge_cells(start_row = debutappreciation+1,start_column = 6, end_row = finappreciation, end_column = 6)
                j=8
                debutjure=8
                for jure in self.jures:
                    if not isinstance(jure, Etudiant) :
                        ws.merge_cells(start_row = debutappreciation,start_column = j+1, end_row = finappreciation, end_column = j+1)
                        ws.cell(row=debutappreciation,column=j).value="=Jury_"+jure.nom.replace(" ","")+'_'+jure.prenom.replace(" ","")+"!"+indextoletter(2*numeroprojet+8)+str(debutappreciation)
                        ws.cell(row=debutappreciation,column=j+1).value="=Jury_"+jure.nom.replace(" ","")+'_'+jure.prenom.replace(" ","")+"!"+indextoletter(1+2*numeroprojet+8)+str(debutappreciation)
                        j+=2
                finjure=j
                ws.cell(row=debutappreciation, column=6).value= "=AVERAGE("+indextoletter(debutjure)+str(debutappreciation)+":"+indextoletter(finjure-1)+str(debutappreciation)+")"
                wssyntheseprojet.cell(row=numeroappreciation+4, column=numeroprojet+2).value="=Projet_"+projet.nom.replace(" ","")+"!F"+str(indiceslignenote[numeroappreciation])
                j=4
                for etudiant in self.jures:
                    if isinstance(etudiant,Etudiant):
                        wssyntheseetudiant.cell(row=j,column=numeroappreciation+4).value="=Projet_"+self.chercher_projet_by_id(etudiant.idprojet).nom.replace(" ","")+"!F"+str(indiceslignenote[numeroappreciation])
                        j+=1
                
                i+=1
                numeroappreciation+=1
            j=8
            numerojure=0
            for jure in self.jures:
                if not isinstance(jure, Etudiant) :
                    ws.merge_cells(start_row = 2,start_column = j, end_row = 2, end_column = j+1)
                    ws.column_dimensions[indextoletter(j)].width = self.canevas.widthNoteJure
                    ws.column_dimensions[indextoletter(j+1)].width = self.canevas.widthCommentaireJure
                    ws.cell(row=2, column=j).value = jure.nom + ' ' + jure.prenom
                    #ws.cell(row = 2, column = j).border = Border(self.canevas.border.left, self.canevas.border.right, self.canevas.border.top, self.canevas.border.bottom)
                    #ws.cell(row = 2, column = j+1).border = Border(self.canevas.border.left, self.canevas.border.right, self.canevas.border.top, self.canevas.border.bottom)
                    ws.cell(row =2, column = j).alignment = Alignment(wrap_text=True)
                    ws.cell(row=2,column=j).font=Font(name=self.canevas.fontTitreJure.name, charset=self.canevas.fontTitreJure.charset, family=self.canevas.fontTitreJure.family, b=self.canevas.fontTitreJure.b, i=self.canevas.fontTitreJure.i, strike=self.canevas.fontTitreJure.strike, outline=self.canevas.fontTitreJure.outline, shadow=self.canevas.fontTitreJure.shadow, condense=self.canevas.fontTitreJure.condense, color=self.canevas.fontTitreJure.color, extend=self.canevas.fontTitreJure.extend, sz=self.canevas.fontTitreJure.size, u=self.canevas.fontTitreJure.u, vertAlign=self.canevas.fontTitreJure.vertAlign, scheme=self.canevas.fontTitreJure.scheme)
                    ws.cell(row=3, column=j).value = self.canevas.titreColonneNoteJure
                    #ws.cell(row = 3, column = j).border = Border(self.canevas.border.left, self.canevas.border.right, self.canevas.border.top, self.canevas.border.bottom)
                    ws.cell(row =3, column = j).alignment = Alignment(wrap_text=True)
                    ws.cell(row=3,column=j).font=Font(name=self.canevas.fontColonneNoteJure.name, charset=self.canevas.fontColonneNoteJure.charset, family=self.canevas.fontColonneNoteJure.family, b=self.canevas.fontColonneNoteJure.b, i=self.canevas.fontColonneNoteJure.i, strike=self.canevas.fontColonneNoteJure.strike, outline=self.canevas.fontColonneNoteJure.outline, shadow=self.canevas.fontColonneNoteJure.shadow, condense=self.canevas.fontColonneNoteJure.condense, color=self.canevas.fontColonneNoteJure.color, extend=self.canevas.fontColonneNoteJure.extend, sz=self.canevas.fontColonneNoteJure.size, u=self.canevas.fontColonneNoteJure.u, vertAlign=self.canevas.fontColonneNoteJure.vertAlign, scheme=self.canevas.fontColonneNoteJure.scheme)
                    ws.cell(row=3, column=j+1).value = self.canevas.titreColonneCommentaireJure
                    #ws.cell(row = 3, column = j+1).border = Border(self.canevas.border.left, self.canevas.border.right, self.canevas.border.top, self.canevas.border.bottom)
                    ws.cell(row =3, column = j+1).alignment = Alignment(wrap_text=True)
                    ws.cell(row=3,column=j+1).font=Font(name=self.canevas.fontColonneCommentaireJure.name, charset=self.canevas.fontColonneCommentaireJure.charset, family=self.canevas.fontColonneCommentaireJure.family, b=self.canevas.fontColonneCommentaireJure.b, i=self.canevas.fontColonneCommentaireJure.i, strike=self.canevas.fontColonneCommentaireJure.strike, outline=self.canevas.fontColonneCommentaireJure.outline, shadow=self.canevas.fontColonneCommentaireJure.shadow, condense=self.canevas.fontColonneCommentaireJure.condense, color=self.canevas.fontColonneCommentaireJure.color, extend=self.canevas.fontColonneCommentaireJure.extend, sz=self.canevas.fontColonneCommentaireJure.size, u=self.canevas.fontColonneCommentaireJure.u, vertAlign=self.canevas.fontColonneCommentaireJure.vertAlign, scheme=self.canevas.fontColonneCommentaireJure.scheme)
                    ws.cell(row=2,column = j).fill = PatternFill(patternType=self.canevas.colors[3+numerojure%2].patternType, fgColor=self.canevas.colors[3+numerojure%2].fgColor, bgColor=self.canevas.colors[3+numerojure%2].bgColor)
                    ws.cell(row=3,column = j).fill = PatternFill(patternType=self.canevas.colors[3+numerojure%2].patternType, fgColor=self.canevas.colors[3+numerojure%2].fgColor, bgColor=self.canevas.colors[3+numerojure%2].bgColor)
                    ws.cell(row=3,column = j+1).fill = PatternFill(patternType=self.canevas.colors[3+numerojure%2].patternType, fgColor=self.canevas.colors[3+numerojure%2].fgColor, bgColor=self.canevas.colors[3+numerojure%2].bgColor)
                    numerojure+=1
                    j+=2
            numeroprojet +=1
        self.wbout.save("FichierSortie_parJury.xlsx")
    
    def generate_jury_sheets(self):
        for jure in self.jures:
            if not isinstance(jure, Etudiant) :
                ws = self.wbout.create_sheet("Jury_"+jure.nom.replace(" ","")+ '_' + jure.prenom.replace(" ",""))
                ws.merge_cells(start_row = 1,start_column = 1, end_row = 1, end_column = 4)
                ws.row_dimensions[2].height=self.canevas.heightJure
                ws.column_dimensions['A'].width=self.canevas.widthMarge1
                ws.column_dimensions['B'].width=self.canevas.widthCritere
                ws.column_dimensions['C'].width=self.canevas.widthPoints
                ws.column_dimensions['D'].width=self.canevas.widthPourcentage
                ws.column_dimensions['E'].width=self.canevas.widthMarge2
                ws.column_dimensions['F'].width=self.canevas.widthNoteFinale
                ws.column_dimensions['G'].width=self.canevas.widthMarge3
                ws.cell(row=1,column=1).value= "Jury_"+jure.nom + '_' + jure.prenom 
                ws.cell(row =1, column = 1).alignment = Alignment(wrap_text=True)
                ws.cell(row=1,column=1).font=Font(name=self.canevas.fontTitre.name, charset=self.canevas.fontTitre.charset, family=self.canevas.fontTitre.family, b=self.canevas.fontTitre.b, i=self.canevas.fontTitre.i, strike=self.canevas.fontTitre.strike, outline=self.canevas.fontTitre.outline, shadow=self.canevas.fontTitre.shadow, condense=self.canevas.fontTitre.condense, color=self.canevas.fontTitre.color, extend=self.canevas.fontTitre.extend, sz=self.canevas.fontTitre.size, u=self.canevas.fontTitre.u, vertAlign=self.canevas.fontTitre.vertAlign, scheme=self.canevas.fontTitre.scheme)
                ws.cell(row=1,column=1).font=Font(name=self.canevas.fontTitre.name, charset=self.canevas.fontTitre.charset, family=self.canevas.fontTitre.family, b=self.canevas.fontTitre.b, i=self.canevas.fontTitre.i, strike=self.canevas.fontTitre.strike, outline=self.canevas.fontTitre.outline, shadow=self.canevas.fontTitre.shadow, condense=self.canevas.fontTitre.condense, color=self.canevas.fontTitre.color, extend=self.canevas.fontTitre.extend, sz=self.canevas.fontTitre.size, u=self.canevas.fontTitre.u, vertAlign=self.canevas.fontTitre.vertAlign, scheme=self.canevas.fontTitre.scheme)
                ws.cell(row=2,column = 4).value = self.canevas.titrePourcentage
                ws.cell(row =2, column = 4).alignment = Alignment(wrap_text=True)
                ws.cell(row=2,column = 4).fill = PatternFill(patternType=self.canevas.colors[1].patternType, fgColor=self.canevas.colors[1].fgColor, bgColor=self.canevas.colors[1].bgColor) 
                ws.cell(row=2,column=4).font=Font(name=self.canevas.fontPourcentage.name, charset=self.canevas.fontPourcentage.charset, family=self.canevas.fontPourcentage.family, b=self.canevas.fontPourcentage.b, i=self.canevas.fontPourcentage.i, strike=self.canevas.fontPourcentage.strike, outline=self.canevas.fontPourcentage.outline, shadow=self.canevas.fontPourcentage.shadow, condense=self.canevas.fontPourcentage.condense, color=self.canevas.fontPourcentage.color, extend=self.canevas.fontPourcentage.extend, sz=self.canevas.fontPourcentage.size, u=self.canevas.fontPourcentage.u, vertAlign=self.canevas.fontPourcentage.vertAlign, scheme=self.canevas.fontPourcentage.scheme)
                ws.cell(row=2,column = 4).fill = PatternFill(patternType=self.canevas.colors[1].patternType, fgColor=self.canevas.colors[1].fgColor, bgColor=self.canevas.colors[1].bgColor) 
                ws.cell(row=3,column = 6).value = self.canevas.titreNoteFinale
                ws.cell(row =3, column = 6).alignment = Alignment(wrap_text=True)
                ws.cell(row=3,column=6).font=Font(name=self.canevas.fontTitreNoteFinale.name, charset=self.canevas.fontTitreNoteFinale.charset, family=self.canevas.fontTitreNoteFinale.family, b=self.canevas.fontTitreNoteFinale.b, i=self.canevas.fontTitreNoteFinale.i, strike=self.canevas.fontTitreNoteFinale.strike, outline=self.canevas.fontTitreNoteFinale.outline, shadow=self.canevas.fontTitreNoteFinale.shadow, condense=self.canevas.fontTitreNoteFinale.condense, color=self.canevas.fontTitreNoteFinale.color, extend=self.canevas.fontTitreNoteFinale.extend, sz=self.canevas.fontTitreNoteFinale.size, u=self.canevas.fontTitreNoteFinale.u, vertAlign=self.canevas.fontTitreNoteFinale.vertAlign, scheme=self.canevas.fontTitreNoteFinale.scheme)
                i=4
                for appreciation in self.canevas.appreciations:
                    debutappreciation=i
                    #ws.cell(row = debutappreciation, column = 6).border = Border(self.canevas.border.left, self.canevas.border.right, self.canevas.border.top, self.canevas.border.bottom)
                    ws.cell(row =debutappreciation, column = 6).alignment = Alignment(wrap_text=True)
                    ws.cell(row=debutappreciation,column = 6).fill = PatternFill(patternType=self.canevas.colors[0].patternType, fgColor=self.canevas.colors[0].fgColor, bgColor=self.canevas.colors[0].bgColor)
                    ws.cell(row=debutappreciation,column=6).font=Font(name=self.canevas.fontNoteFinale.name, charset=self.canevas.fontNoteFinale.charset, family=self.canevas.fontNoteFinale.family, b=self.canevas.fontNoteFinale.b, i=self.canevas.fontNoteFinale.i, strike=self.canevas.fontNoteFinale.strike, outline=self.canevas.fontNoteFinale.outline, shadow=self.canevas.fontNoteFinale.shadow, condense=self.canevas.fontNoteFinale.condense, color=self.canevas.fontNoteFinale.color, extend=self.canevas.fontNoteFinale.extend, sz=self.canevas.fontNoteFinale.size, u=self.canevas.fontNoteFinale.u, vertAlign=self.canevas.fontNoteFinale.vertAlign, scheme=self.canevas.fontNoteFinale.scheme)
                    ws.merge_cells(start_row = i,start_column = 1, end_row = i, end_column = 2)
                    ws.cell(row=i,column = 1).value = appreciation.titre
                    ws.cell(row =i, column = 1).alignment = Alignment(wrap_text=True)
                    ws.cell(row=i,column=1).font=Font(name=self.canevas.fontAppreciation.name, charset=self.canevas.fontAppreciation.charset, family=self.canevas.fontAppreciation.family, b=self.canevas.fontAppreciation.b, i=self.canevas.fontAppreciation.i, strike=self.canevas.fontAppreciation.strike, outline=self.canevas.fontAppreciation.outline, shadow=self.canevas.fontAppreciation.shadow, condense=self.canevas.fontAppreciation.condense, color=self.canevas.fontAppreciation.color, extend=self.canevas.fontAppreciation.extend, sz=self.canevas.fontAppreciation.size, u=self.canevas.fontAppreciation.u, vertAlign=self.canevas.fontAppreciation.vertAlign, scheme=self.canevas.fontAppreciation.scheme)
                    ws.cell(row=i,column = 1).fill = PatternFill(patternType=self.canevas.colors[0].patternType, fgColor=self.canevas.colors[0].fgColor, bgColor=self.canevas.colors[0].bgColor)
                    ws.cell(row=i,column = 3).value = appreciation.nbPoints
                    ws.cell(row =i, column = 3).alignment = Alignment(wrap_text=True)
                    ws.cell(row=i,column=3).font=Font(name=self.canevas.fontNoteAppreciation.name, charset=self.canevas.fontNoteAppreciation.charset, family=self.canevas.fontNoteAppreciation.family, b=self.canevas.fontNoteAppreciation.b, i=self.canevas.fontNoteAppreciation.i, strike=self.canevas.fontNoteAppreciation.strike, outline=self.canevas.fontNoteAppreciation.outline, shadow=self.canevas.fontNoteAppreciation.shadow, condense=self.canevas.fontNoteAppreciation.condense, color=self.canevas.fontNoteAppreciation.color, extend=self.canevas.fontNoteAppreciation.extend, sz=self.canevas.fontNoteAppreciation.size, u=self.canevas.fontNoteAppreciation.u, vertAlign=self.canevas.fontNoteAppreciation.vertAlign, scheme=self.canevas.fontNoteAppreciation.scheme)
                    ws.cell(row=i,column = 3).fill = PatternFill(patternType=self.canevas.colors[0].patternType, fgColor=self.canevas.colors[0].fgColor, bgColor=self.canevas.colors[0].bgColor) 
                    ws.cell(row=i,column = 4).value = str(int(float(appreciation.partDeLaNote)*100)) + '%'
                    ws.cell(row =i, column = 4).alignment = Alignment(wrap_text=True) 
                    ws.cell(row=i,column=4).font=Font(name=self.canevas.fontNotePourcentage.name, charset=self.canevas.fontNotePourcentage.charset, family=self.canevas.fontNotePourcentage.family, b=self.canevas.fontNotePourcentage.b, i=self.canevas.fontNotePourcentage.i, strike=self.canevas.fontNotePourcentage.strike, outline=self.canevas.fontNotePourcentage.outline, shadow=self.canevas.fontNotePourcentage.shadow, condense=self.canevas.fontNotePourcentage.condense, color=self.canevas.fontNotePourcentage.color, extend=self.canevas.fontNotePourcentage.extend, sz=self.canevas.fontNotePourcentage.size, u=self.canevas.fontNotePourcentage.u, vertAlign=self.canevas.fontNotePourcentage.vertAlign, scheme=self.canevas.fontNotePourcentage.scheme)
                    ws.cell(row=i,column = 4).fill = PatternFill(patternType=self.canevas.colors[1].patternType, fgColor=self.canevas.colors[1].fgColor, bgColor=self.canevas.colors[1].bgColor) 
                    i+=1
                    for critere in appreciation.criteres:
                        ws.merge_cells(start_row = i,start_column = 1, end_row =i, end_column = 2)
                        ws.cell(row=i,column = 1).value = critere.titre
                        ws.cell(row =i, column = 1).alignment = Alignment(wrap_text=True)
                        ws.cell(row=i,column=1).font=Font(name=self.canevas.fontTitreCritere.name, charset=self.canevas.fontTitreCritere.charset, family=self.canevas.fontTitreCritere.family, b=self.canevas.fontTitreCritere.b, i=self.canevas.fontTitreCritere.i, strike=self.canevas.fontTitreCritere.strike, outline=self.canevas.fontTitreCritere.outline, shadow=self.canevas.fontTitreCritere.shadow, condense=self.canevas.fontTitreCritere.condense, color=self.canevas.fontTitreCritere.color, extend=self.canevas.fontTitreCritere.extend, sz=self.canevas.fontTitreCritere.size, u=self.canevas.fontTitreCritere.u, vertAlign=self.canevas.fontTitreCritere.vertAlign, scheme=self.canevas.fontTitreCritere.scheme)
                        ws.cell(row=i,column = 3).value = critere.points
                        ws.cell(row =i, column = 3).alignment = Alignment(wrap_text=True)
                        ws.cell(row=i,column=3).font=Font(name=self.canevas.fontNoteCritere.name, charset=self.canevas.fontNoteCritere.charset, family=self.canevas.fontNoteCritere.family, b=self.canevas.fontNoteCritere.b, i=self.canevas.fontNoteCritere.i, strike=self.canevas.fontNoteCritere.strike, outline=self.canevas.fontNoteCritere.outline, shadow=self.canevas.fontNoteCritere.shadow, condense=self.canevas.fontNoteCritere.condense, color=self.canevas.fontNoteCritere.color, extend=self.canevas.fontNoteCritere.extend, sz=self.canevas.fontNoteCritere.size, u=self.canevas.fontNoteCritere.u, vertAlign=self.canevas.fontNoteCritere.vertAlign, scheme=self.canevas.fontNoteCritere.scheme)
                        if (critere.description == None):
                            i+=1
                        else:
                            j=8 #indice de colonnes parcourues à travers les jurés
                            for projet in self.projets:
                                ws.merge_cells(start_row = i,start_column = j, end_row = i+1, end_column = j)
                                j+=2
                            ws.merge_cells(start_row = i,start_column = 3, end_row = i+1, end_column = 3)
                            ws.row_dimensions[i+1].height = critere.hauteur
                            ws.cell(row=i+1,column = 2).value = critere.description
                            ws.cell(row =i+1, column = 2).alignment = Alignment(wrap_text=True)
                            ws.cell(row=i+1,column=2).font=Font(name=self.canevas.fontDescriptionCritere.name, charset=self.canevas.fontDescriptionCritere.charset, family=self.canevas.fontDescriptionCritere.family, b=self.canevas.fontDescriptionCritere.b, i=self.canevas.fontDescriptionCritere.i, strike=self.canevas.fontDescriptionCritere.strike, outline=self.canevas.fontDescriptionCritere.outline, shadow=self.canevas.fontDescriptionCritere.shadow, condense=self.canevas.fontDescriptionCritere.condense, color=self.canevas.fontDescriptionCritere.color, extend=self.canevas.fontDescriptionCritere.extend, sz=self.canevas.fontDescriptionCritere.size, u=self.canevas.fontDescriptionCritere.u, vertAlign=self.canevas.fontDescriptionCritere.vertAlign, scheme=self.canevas.fontDescriptionCritere.scheme)
                            i+=2
                    finappreciation=i-1
                    ws.merge_cells(start_row = debutappreciation+1,start_column = 6, end_row = finappreciation, end_column = 6)
                    j=8
                    debutprojet = j
                    for projet in self.projets:
                        ws.cell(row=debutappreciation,column=j).value='=SUM('+indextoletter(j)+str(debutappreciation+1)+':'+indextoletter(j)+str(finappreciation)+')'
                        ws.merge_cells(start_row = debutappreciation,start_column = j+1, end_row = finappreciation, end_column = j+1)
                        j+=2
                    finprojet=j
                    ws.cell(row=debutappreciation, column=6).value= "=AVERAGE("+indextoletter(debutprojet)+str(debutappreciation)+":"+indextoletter(finprojet-1)+str(debutappreciation)+")"
                    i+=1
                j=8
                numeroprojet=0
                for projet in self.projets:
                    ws.merge_cells(start_row = 2,start_column = j, end_row = 2, end_column = j+1)
                    ws.column_dimensions[indextoletter(j)].width = self.canevas.widthNoteJure
                    ws.column_dimensions[indextoletter(j+1)].width = self.canevas.widthCommentaireJure
                    ws.cell(row=2, column=j).value = projet.nom
                    ws.cell(row =2, column = j).alignment = Alignment(wrap_text=True)
                    ws.cell(row=2,column=j).font=Font(name=self.canevas.fontTitreJure.name, charset=self.canevas.fontTitreJure.charset, family=self.canevas.fontTitreJure.family, b=self.canevas.fontTitreJure.b, i=self.canevas.fontTitreJure.i, strike=self.canevas.fontTitreJure.strike, outline=self.canevas.fontTitreJure.outline, shadow=self.canevas.fontTitreJure.shadow, condense=self.canevas.fontTitreJure.condense, color=self.canevas.fontTitreJure.color, extend=self.canevas.fontTitreJure.extend, sz=self.canevas.fontTitreJure.size, u=self.canevas.fontTitreJure.u, vertAlign=self.canevas.fontTitreJure.vertAlign, scheme=self.canevas.fontTitreJure.scheme)
                    ws.cell(row=3, column=j).value = self.canevas.titreColonneNoteJure
                    ws.cell(row =3, column = j).alignment = Alignment(wrap_text=True)
                    ws.cell(row=3,column=j).font=Font(name=self.canevas.fontColonneNoteJure.name, charset=self.canevas.fontColonneNoteJure.charset, family=self.canevas.fontColonneNoteJure.family, b=self.canevas.fontColonneNoteJure.b, i=self.canevas.fontColonneNoteJure.i, strike=self.canevas.fontColonneNoteJure.strike, outline=self.canevas.fontColonneNoteJure.outline, shadow=self.canevas.fontColonneNoteJure.shadow, condense=self.canevas.fontColonneNoteJure.condense, color=self.canevas.fontColonneNoteJure.color, extend=self.canevas.fontColonneNoteJure.extend, sz=self.canevas.fontColonneNoteJure.size, u=self.canevas.fontColonneNoteJure.u, vertAlign=self.canevas.fontColonneNoteJure.vertAlign, scheme=self.canevas.fontColonneNoteJure.scheme)
                    ws.cell(row=3, column=j+1).value = self.canevas.titreColonneCommentaireJure
                    ws.cell(row =3, column = j+1).alignment = Alignment(wrap_text=True)
                    ws.cell(row=3,column=j+1).font=Font(name=self.canevas.fontColonneCommentaireJure.name, charset=self.canevas.fontColonneCommentaireJure.charset, family=self.canevas.fontColonneCommentaireJure.family, b=self.canevas.fontColonneCommentaireJure.b, i=self.canevas.fontColonneCommentaireJure.i, strike=self.canevas.fontColonneCommentaireJure.strike, outline=self.canevas.fontColonneCommentaireJure.outline, shadow=self.canevas.fontColonneCommentaireJure.shadow, condense=self.canevas.fontColonneCommentaireJure.condense, color=self.canevas.fontColonneCommentaireJure.color, extend=self.canevas.fontColonneCommentaireJure.extend, sz=self.canevas.fontColonneCommentaireJure.size, u=self.canevas.fontColonneCommentaireJure.u, vertAlign=self.canevas.fontColonneCommentaireJure.vertAlign, scheme=self.canevas.fontColonneCommentaireJure.scheme)
                    ws.cell(row=2,column = j).fill = PatternFill(patternType=self.canevas.colors[3+numeroprojet%2].patternType, fgColor=self.canevas.colors[3+numeroprojet%2].fgColor, bgColor=self.canevas.colors[3+numeroprojet%2].bgColor)
                    ws.cell(row=3,column = j).fill = PatternFill(patternType=self.canevas.colors[3+numeroprojet%2].patternType, fgColor=self.canevas.colors[3+numeroprojet%2].fgColor, bgColor=self.canevas.colors[3+numeroprojet%2].bgColor)
                    ws.cell(row=3,column = j+1).fill = PatternFill(patternType=self.canevas.colors[3+numeroprojet%2].patternType, fgColor=self.canevas.colors[3+numeroprojet%2].fgColor, bgColor=self.canevas.colors[3+numeroprojet%2].bgColor)
                    numeroprojet+=1
                    j+=2
                
            
            
            
        self.wbout.save("FichierSortie_parJury.xlsx")
        
        
    def generate_synthese_projet(self):
        width_columns_projets = 21
        ws = self.wbout.create_sheet("Synthese_projet")
        ws.merge_cells(start_row = 1,start_column = 1, end_row = 1, end_column = 4)
        ws.cell(row =1, column = 1).value =  "Synthese notes projets"
        ws.cell(row =1, column = 1).alignment = Alignment(wrap_text=True)
        ws.column_dimensions['A'].width = 34
        j=4
        debutappreciation=j
        for appreciation in self.canevas.appreciations: 
            ws.cell(row=j, column= 1).value =appreciation.titre
            ws.cell(row =j, column = 1).alignment = Alignment(wrap_text=True)
            j+=1
        finappreciation=j
        for i in range(debutappreciation-1,finappreciation) :
            ws.cell(row = i, column = 1).border = Border(self.canevas.border.left, self.canevas.border.right, self.canevas.border.top, self.canevas.border.bottom)
        j = 2
        for projet in self.projets:
            ws.column_dimensions[indextoletter(j)].width = width_columns_projets
            ws.cell(row = 3, column = j).value = self.projets[j-2].nom
            ws.cell(row =3, column = j).alignment = Alignment(wrap_text=True)
            ws.cell(row = 3, column = j).font = Font(bold=True)
            for i in range(debutappreciation-1,finappreciation) :
                ws.cell(row = i, column = j).border = Border(self.canevas.border.left, self.canevas.border.right, self.canevas.border.top, self.canevas.border.bottom)
            j += 1
        self.wbout.save("FichierSortie_parJury.xlsx")   
        
        
    def generate_sortie_jury(self):
        self.wbout = Workbook()
        self.generate_synthese_etudiant()
        self.wbout.remove(self.wbout[self.wbout.sheetnames[0]])
        self.wbout.save("FichierSortie_parJury.xlsx")  
        self.generate_synthese_projet()
        self.generate_jury_sheets()
        self.generate_projets_sheets()
        
        
       

## VI - Génération de la feuille sortie par Etudiant
        
        
    def generate_synthese_projet_etudiant(self):
        width_columns_projets = 21
        ws = self.wbout.create_sheet("Synthese_projet")
        ws.merge_cells(start_row = 1,start_column = 1, end_row = 1, end_column = 4)
        ws.cell(row =1, column = 1).value =  "Synthese notes projets"
        ws.cell(row =1, column = 1).alignment = Alignment(wrap_text=True)
        ws.column_dimensions['A'].width = 34
        j=4
        debutappreciation=j
        for appreciation in self.canevas.appreciations: 
            ws.cell(row=j, column= 1).value =appreciation.titre
            ws.cell(row =j, column = 1).alignment = Alignment(wrap_text=True)
            j+=1
        finappreciation=j
        for i in range(debutappreciation-1,finappreciation) :
            ws.cell(row = i, column = 1).border = Border(self.canevas.border.left, self.canevas.border.right, self.canevas.border.top, self.canevas.border.bottom)
        j = 2
        for projet in self.projets:
            ws.column_dimensions[indextoletter(j)].width = width_columns_projets
            ws.cell(row = 3, column = j).value = self.projets[j-2].nom
            ws.cell(row =3, column = j).alignment = Alignment(wrap_text=True)
            ws.cell(row = 3, column = j).font = Font(bold=True)
            for i in range(debutappreciation-1,finappreciation) :
                ws.cell(row = i, column = j).border = Border(self.canevas.border.left, self.canevas.border.right, self.canevas.border.top, self.canevas.border.bottom)
            j += 1
        
        self.wbout.save("FichierSortie_parEtudiant.xlsx")
    
    
    def generate_projets_sheets_etudiant(self):
        wssyntheseprojet=self.wbout[self.wbout.sheetnames[0]]
        numeroprojet=0
        for projet in self.projets:
            ws = self.wbout.create_sheet("Projet_"+projet.nom.replace(" ",""))
            ws.merge_cells(start_row = 1,start_column = 1, end_row = 1, end_column = 4)
            ws.row_dimensions[2].height=self.canevas.heightJure
            ws.column_dimensions['A'].width=self.canevas.widthMarge1
            ws.column_dimensions['B'].width=self.canevas.widthCritere
            ws.column_dimensions['C'].width=self.canevas.widthPoints
            ws.column_dimensions['D'].width=self.canevas.widthPourcentage
            ws.column_dimensions['E'].width=self.canevas.widthMarge2
            ws.column_dimensions['F'].width=self.canevas.widthNoteFinale
            ws.column_dimensions['G'].width=self.canevas.widthMarge3
            ws.cell(row=1,column=1).value= "Projet_"+projet.nom
            ws.cell(row =1, column = 1).alignment = Alignment(wrap_text=True)
            ws.cell(row=1,column=1).font=Font(name=self.canevas.fontTitre.name, charset=self.canevas.fontTitre.charset, family=self.canevas.fontTitre.family, b=self.canevas.fontTitre.b, i=self.canevas.fontTitre.i, strike=self.canevas.fontTitre.strike, outline=self.canevas.fontTitre.outline, shadow=self.canevas.fontTitre.shadow, condense=self.canevas.fontTitre.condense, color=self.canevas.fontTitre.color, extend=self.canevas.fontTitre.extend, sz=self.canevas.fontTitre.size, u=self.canevas.fontTitre.u, vertAlign=self.canevas.fontTitre.vertAlign, scheme=self.canevas.fontTitre.scheme)
            ws.cell(row=2,column = 4).value = self.canevas.titrePourcentage
            ws.cell(row =2, column = 4).alignment = Alignment(wrap_text=True)
            ws.cell(row=2,column = 4).fill = PatternFill(patternType=self.canevas.colors[1].patternType, fgColor=self.canevas.colors[1].fgColor, bgColor=self.canevas.colors[1].bgColor) 
            ws.cell(row=2,column=4).font=Font(name=self.canevas.fontPourcentage.name, charset=self.canevas.fontPourcentage.charset, family=self.canevas.fontPourcentage.family, b=self.canevas.fontPourcentage.b, i=self.canevas.fontPourcentage.i, strike=self.canevas.fontPourcentage.strike, outline=self.canevas.fontPourcentage.outline, shadow=self.canevas.fontPourcentage.shadow, condense=self.canevas.fontPourcentage.condense, color=self.canevas.fontPourcentage.color, extend=self.canevas.fontPourcentage.extend, sz=self.canevas.fontPourcentage.size, u=self.canevas.fontPourcentage.u, vertAlign=self.canevas.fontPourcentage.vertAlign, scheme=self.canevas.fontPourcentage.scheme)
            ws.cell(row=3,column = 6).value = self.canevas.titreNoteFinale
            ws.cell(row =3, column = 6).alignment = Alignment(wrap_text=True)
            ws.cell(row=3,column=6).font=Font(name=self.canevas.fontTitreNoteFinale.name, charset=self.canevas.fontTitreNoteFinale.charset, family=self.canevas.fontTitreNoteFinale.family, b=self.canevas.fontTitreNoteFinale.b, i=self.canevas.fontTitreNoteFinale.i, strike=self.canevas.fontTitreNoteFinale.strike, outline=self.canevas.fontTitreNoteFinale.outline, shadow=self.canevas.fontTitreNoteFinale.shadow, condense=self.canevas.fontTitreNoteFinale.condense, color=self.canevas.fontTitreNoteFinale.color, extend=self.canevas.fontTitreNoteFinale.extend, sz=self.canevas.fontTitreNoteFinale.size, u=self.canevas.fontTitreNoteFinale.u, vertAlign=self.canevas.fontTitreNoteFinale.vertAlign, scheme=self.canevas.fontTitreNoteFinale.scheme)
            i=4
            numeroappreciation=0
            indiceslignenote=[]
            for appreciation in self.canevas.appreciations:
                debutappreciation=i
                indiceslignenote.append(debutappreciation)
                #ws.cell(row = debutappreciation, column = 6).border = Border(self.canevas.border.left, self.canevas.border.right, self.canevas.border.top, self.canevas.border.bottom)
                ws.cell(row =debutappreciation, column = 6).alignment = Alignment(wrap_text=True)
                ws.cell(row=debutappreciation,column = 6).fill = PatternFill(patternType=self.canevas.colors[0].patternType, fgColor=self.canevas.colors[0].fgColor, bgColor=self.canevas.colors[0].bgColor)
                ws.cell(row=debutappreciation,column=6).font=Font(name=self.canevas.fontNoteFinale.name, charset=self.canevas.fontNoteFinale.charset, family=self.canevas.fontNoteFinale.family, b=self.canevas.fontNoteFinale.b, i=self.canevas.fontNoteFinale.i, strike=self.canevas.fontNoteFinale.strike, outline=self.canevas.fontNoteFinale.outline, shadow=self.canevas.fontNoteFinale.shadow, condense=self.canevas.fontNoteFinale.condense, color=self.canevas.fontNoteFinale.color, extend=self.canevas.fontNoteFinale.extend, sz=self.canevas.fontNoteFinale.size, u=self.canevas.fontNoteFinale.u, vertAlign=self.canevas.fontNoteFinale.vertAlign, scheme=self.canevas.fontNoteFinale.scheme)
                ws.merge_cells(start_row = i,start_column = 1, end_row = i, end_column = 2)
                ws.cell(row=i,column = 1).value = appreciation.titre
                ws.cell(row =i, column = 1).alignment = Alignment(wrap_text=True)
                ws.cell(row=i,column=1).font=Font(name=self.canevas.fontAppreciation.name, charset=self.canevas.fontAppreciation.charset, family=self.canevas.fontAppreciation.family, b=self.canevas.fontAppreciation.b, i=self.canevas.fontAppreciation.i, strike=self.canevas.fontAppreciation.strike, outline=self.canevas.fontAppreciation.outline, shadow=self.canevas.fontAppreciation.shadow, condense=self.canevas.fontAppreciation.condense, color=self.canevas.fontAppreciation.color, extend=self.canevas.fontAppreciation.extend, sz=self.canevas.fontAppreciation.size, u=self.canevas.fontAppreciation.u, vertAlign=self.canevas.fontAppreciation.vertAlign, scheme=self.canevas.fontAppreciation.scheme)
                ws.cell(row=i,column = 1).fill = PatternFill(patternType=self.canevas.colors[0].patternType, fgColor=self.canevas.colors[0].fgColor, bgColor=self.canevas.colors[0].bgColor)
                ws.cell(row=i,column = 3).value = appreciation.nbPoints
                ws.cell(row =i, column = 3).alignment = Alignment(wrap_text=True)
                ws.cell(row=i,column=3).font=Font(name=self.canevas.fontNoteAppreciation.name, charset=self.canevas.fontNoteAppreciation.charset, family=self.canevas.fontNoteAppreciation.family, b=self.canevas.fontNoteAppreciation.b, i=self.canevas.fontNoteAppreciation.i, strike=self.canevas.fontNoteAppreciation.strike, outline=self.canevas.fontNoteAppreciation.outline, shadow=self.canevas.fontNoteAppreciation.shadow, condense=self.canevas.fontNoteAppreciation.condense, color=self.canevas.fontNoteAppreciation.color, extend=self.canevas.fontNoteAppreciation.extend, sz=self.canevas.fontNoteAppreciation.size, u=self.canevas.fontNoteAppreciation.u, vertAlign=self.canevas.fontNoteAppreciation.vertAlign, scheme=self.canevas.fontNoteAppreciation.scheme)
                ws.cell(row=i,column = 3).fill = PatternFill(patternType=self.canevas.colors[0].patternType, fgColor=self.canevas.colors[0].fgColor, bgColor=self.canevas.colors[0].bgColor) 
                ws.cell(row=i,column = 4).value = str(int(float(appreciation.partDeLaNote)*100)) + '%'
                ws.cell(row =i, column = 4).alignment = Alignment(wrap_text=True) 
                ws.cell(row=i,column=4).font=Font(name=self.canevas.fontNotePourcentage.name, charset=self.canevas.fontNotePourcentage.charset, family=self.canevas.fontNotePourcentage.family, b=self.canevas.fontNotePourcentage.b, i=self.canevas.fontNotePourcentage.i, strike=self.canevas.fontNotePourcentage.strike, outline=self.canevas.fontNotePourcentage.outline, shadow=self.canevas.fontNotePourcentage.shadow, condense=self.canevas.fontNotePourcentage.condense, color=self.canevas.fontNotePourcentage.color, extend=self.canevas.fontNotePourcentage.extend, sz=self.canevas.fontNotePourcentage.size, u=self.canevas.fontNotePourcentage.u, vertAlign=self.canevas.fontNotePourcentage.vertAlign, scheme=self.canevas.fontNotePourcentage.scheme)
                ws.cell(row=i,column = 4).fill = PatternFill(patternType=self.canevas.colors[1].patternType, fgColor=self.canevas.colors[1].fgColor, bgColor=self.canevas.colors[1].bgColor) 
                i+=1
                for critere in appreciation.criteres:
                    ws.merge_cells(start_row = i,start_column = 1, end_row =i, end_column = 2)
                    ws.cell(row=i,column = 1).value = critere.titre
                    ws.cell(row =i, column = 1).alignment = Alignment(wrap_text=True)
                    ws.cell(row=i,column=1).font=Font(name=self.canevas.fontTitreCritere.name, charset=self.canevas.fontTitreCritere.charset, family=self.canevas.fontTitreCritere.family, b=self.canevas.fontTitreCritere.b, i=self.canevas.fontTitreCritere.i, strike=self.canevas.fontTitreCritere.strike, outline=self.canevas.fontTitreCritere.outline, shadow=self.canevas.fontTitreCritere.shadow, condense=self.canevas.fontTitreCritere.condense, color=self.canevas.fontTitreCritere.color, extend=self.canevas.fontTitreCritere.extend, sz=self.canevas.fontTitreCritere.size, u=self.canevas.fontTitreCritere.u, vertAlign=self.canevas.fontTitreCritere.vertAlign, scheme=self.canevas.fontTitreCritere.scheme)
                    ws.cell(row=i,column = 3).value = critere.points
                    ws.cell(row =i, column = 3).alignment = Alignment(wrap_text=True)
                    ws.cell(row=i,column=3).font=Font(name=self.canevas.fontNoteCritere.name, charset=self.canevas.fontNoteCritere.charset, family=self.canevas.fontNoteCritere.family, b=self.canevas.fontNoteCritere.b, i=self.canevas.fontNoteCritere.i, strike=self.canevas.fontNoteCritere.strike, outline=self.canevas.fontNoteCritere.outline, shadow=self.canevas.fontNoteCritere.shadow, condense=self.canevas.fontNoteCritere.condense, color=self.canevas.fontNoteCritere.color, extend=self.canevas.fontNoteCritere.extend, sz=self.canevas.fontNoteCritere.size, u=self.canevas.fontNoteCritere.u, vertAlign=self.canevas.fontNoteCritere.vertAlign, scheme=self.canevas.fontNoteCritere.scheme)
                    j=8
                    debutjure=j
                    for jure in self.jures:
                        if isinstance(jure, Etudiant) :
                            ws.cell(row=i,column=j).value="=Etudiant_"+jure.nom.replace(" ","")+'_'+jure.prenom.replace(" ","")+"!"+indextoletter(2*numeroprojet+8)+str(i)
                            j+=2
                    finjure=j
                    if (critere.description == None):
                        i+=1
                    else:
                        j=8 #indice de colonnes parcourues à travers les jurés
                        for jure in self.jures:
                            if isinstance(jure, Etudiant) :
                                ws.merge_cells(start_row = i,start_column = j, end_row = i+1, end_column = j)
                                j+=2
                        ws.merge_cells(start_row = i,start_column = 3, end_row = i+1, end_column = 3)
                        ws.row_dimensions[i+1].height = critere.hauteur
                        ws.cell(row=i+1,column = 2).value = critere.description
                        ws.cell(row =i+1, column = 2).alignment = Alignment(wrap_text=True)
                        ws.cell(row=i+1,column=2).font=Font(name=self.canevas.fontDescriptionCritere.name, charset=self.canevas.fontDescriptionCritere.charset, family=self.canevas.fontDescriptionCritere.family, b=self.canevas.fontDescriptionCritere.b, i=self.canevas.fontDescriptionCritere.i, strike=self.canevas.fontDescriptionCritere.strike, outline=self.canevas.fontDescriptionCritere.outline, shadow=self.canevas.fontDescriptionCritere.shadow, condense=self.canevas.fontDescriptionCritere.condense, color=self.canevas.fontDescriptionCritere.color, extend=self.canevas.fontDescriptionCritere.extend, sz=self.canevas.fontDescriptionCritere.size, u=self.canevas.fontDescriptionCritere.u, vertAlign=self.canevas.fontDescriptionCritere.vertAlign, scheme=self.canevas.fontDescriptionCritere.scheme)
                        i+=2
                finappreciation=i-1
                ws.merge_cells(start_row = debutappreciation+1,start_column = 6, end_row = finappreciation, end_column = 6)
                j=8
                debutjure=j
                for jure in self.jures:
                    if isinstance(jure, Etudiant) :
                        ws.merge_cells(start_row = debutappreciation,start_column = j+1, end_row = finappreciation, end_column = j+1)
                        ws.cell(row=debutappreciation,column=j).value="=Etudiant_"+jure.nom.replace(" ","")+'_'+jure.prenom.replace(" ","")+"!"+indextoletter(2*numeroprojet+8)+str(debutappreciation)
                        ws.cell(row=debutappreciation,column=j+1).value="=Etudiant_"+jure.nom.replace(" ","")+'_'+jure.prenom.replace(" ","")+"!"+indextoletter(1+2*numeroprojet+8)+str(debutappreciation)
                        j+=2
                finjure=j
                ws.cell(row=debutappreciation, column=6).value= "=AVERAGE("+indextoletter(debutjure)+str(debutappreciation)+":"+indextoletter(finjure-1)+str(debutappreciation)+")"
                wssyntheseprojet.cell(row=numeroappreciation+4, column=numeroprojet+2).value="=Projet_"+projet.nom.replace(" ","")+"!"+"F"+str(indiceslignenote[numeroappreciation])
                i+=1
                numeroappreciation+=1
            j=8
            numerojure=0
            for jure in self.jures:
                if isinstance(jure, Etudiant) :
                    ws.merge_cells(start_row = 2,start_column = j, end_row = 2, end_column = j+1)
                    ws.column_dimensions[indextoletter(j)].width = self.canevas.widthNoteJure
                    ws.column_dimensions[indextoletter(j+1)].width = self.canevas.widthCommentaireJure
                    ws.cell(row=2, column=j).value = jure.nom + ' ' + jure.prenom
                    ws.cell(row =2, column = j).alignment = Alignment(wrap_text=True)
                    ws.cell(row=2,column=j).font=Font(name=self.canevas.fontTitreJure.name, charset=self.canevas.fontTitreJure.charset, family=self.canevas.fontTitreJure.family, b=self.canevas.fontTitreJure.b, i=self.canevas.fontTitreJure.i, strike=self.canevas.fontTitreJure.strike, outline=self.canevas.fontTitreJure.outline, shadow=self.canevas.fontTitreJure.shadow, condense=self.canevas.fontTitreJure.condense, color=self.canevas.fontTitreJure.color, extend=self.canevas.fontTitreJure.extend, sz=self.canevas.fontTitreJure.size, u=self.canevas.fontTitreJure.u, vertAlign=self.canevas.fontTitreJure.vertAlign, scheme=self.canevas.fontTitreJure.scheme)
                    ws.cell(row=3, column=j).value = self.canevas.titreColonneNoteJure
                    ws.cell(row =3, column = j).alignment = Alignment(wrap_text=True)
                    ws.cell(row=3,column=j).font=Font(name=self.canevas.fontColonneNoteJure.name, charset=self.canevas.fontColonneNoteJure.charset, family=self.canevas.fontColonneNoteJure.family, b=self.canevas.fontColonneNoteJure.b, i=self.canevas.fontColonneNoteJure.i, strike=self.canevas.fontColonneNoteJure.strike, outline=self.canevas.fontColonneNoteJure.outline, shadow=self.canevas.fontColonneNoteJure.shadow, condense=self.canevas.fontColonneNoteJure.condense, color=self.canevas.fontColonneNoteJure.color, extend=self.canevas.fontColonneNoteJure.extend, sz=self.canevas.fontColonneNoteJure.size, u=self.canevas.fontColonneNoteJure.u, vertAlign=self.canevas.fontColonneNoteJure.vertAlign, scheme=self.canevas.fontColonneNoteJure.scheme)
                    ws.cell(row=3, column=j+1).value = self.canevas.titreColonneCommentaireJure
                    ws.cell(row =3, column = j+1).alignment = Alignment(wrap_text=True)
                    ws.cell(row=3,column=j+1).font=Font(name=self.canevas.fontColonneCommentaireJure.name, charset=self.canevas.fontColonneCommentaireJure.charset, family=self.canevas.fontColonneCommentaireJure.family, b=self.canevas.fontColonneCommentaireJure.b, i=self.canevas.fontColonneCommentaireJure.i, strike=self.canevas.fontColonneCommentaireJure.strike, outline=self.canevas.fontColonneCommentaireJure.outline, shadow=self.canevas.fontColonneCommentaireJure.shadow, condense=self.canevas.fontColonneCommentaireJure.condense, color=self.canevas.fontColonneCommentaireJure.color, extend=self.canevas.fontColonneCommentaireJure.extend, sz=self.canevas.fontColonneCommentaireJure.size, u=self.canevas.fontColonneCommentaireJure.u, vertAlign=self.canevas.fontColonneCommentaireJure.vertAlign, scheme=self.canevas.fontColonneCommentaireJure.scheme)
                    ws.cell(row=2,column = j).fill = PatternFill(patternType=self.canevas.colors[3+numerojure%2].patternType, fgColor=self.canevas.colors[3+numerojure%2].fgColor, bgColor=self.canevas.colors[3+numerojure%2].bgColor)
                    ws.cell(row=3,column = j).fill = PatternFill(patternType=self.canevas.colors[3+numerojure%2].patternType, fgColor=self.canevas.colors[3+numerojure%2].fgColor, bgColor=self.canevas.colors[3+numerojure%2].bgColor)
                    ws.cell(row=3,column = j+1).fill = PatternFill(patternType=self.canevas.colors[3+numerojure%2].patternType, fgColor=self.canevas.colors[3+numerojure%2].fgColor, bgColor=self.canevas.colors[3+numerojure%2].bgColor)
                    numerojure+=1
                    j+=2
            numeroprojet+=1
        self.wbout.save("FichierSortie_parEtudiant.xlsx")
        
        
    def generate_etudiant_sheets(self):
        for jure in self.jures:
            if isinstance(jure, Etudiant) :
                ws = self.wbout.create_sheet("Etudiant_"+jure.nom.replace(" ","")+ '_' + jure.prenom.replace(" ",""))
                ws.merge_cells(start_row = 1,start_column = 1, end_row = 1, end_column = 4)
                ws.row_dimensions[2].height=self.canevas.heightJure
                ws.column_dimensions['A'].width=self.canevas.widthMarge1
                ws.column_dimensions['B'].width=self.canevas.widthCritere
                ws.column_dimensions['C'].width=self.canevas.widthPoints
                ws.column_dimensions['D'].width=self.canevas.widthPourcentage
                ws.column_dimensions['E'].width=self.canevas.widthMarge2
                ws.column_dimensions['F'].width=self.canevas.widthNoteFinale
                ws.column_dimensions['G'].width=self.canevas.widthMarge3
                ws.cell(row=1,column=1).value= "Etudiant"+jure.nom + '_' + jure.prenom
                ws.cell(row =1, column = 1).alignment = Alignment(wrap_text=True)
                ws.cell(row=1,column=1).font=Font(name=self.canevas.fontTitre.name, charset=self.canevas.fontTitre.charset, family=self.canevas.fontTitre.family, b=self.canevas.fontTitre.b, i=self.canevas.fontTitre.i, strike=self.canevas.fontTitre.strike, outline=self.canevas.fontTitre.outline, shadow=self.canevas.fontTitre.shadow, condense=self.canevas.fontTitre.condense, color=self.canevas.fontTitre.color, extend=self.canevas.fontTitre.extend, sz=self.canevas.fontTitre.size, u=self.canevas.fontTitre.u, vertAlign=self.canevas.fontTitre.vertAlign, scheme=self.canevas.fontTitre.scheme)
                ws.cell(row=1,column=1).font=Font(name=self.canevas.fontTitre.name, charset=self.canevas.fontTitre.charset, family=self.canevas.fontTitre.family, b=self.canevas.fontTitre.b, i=self.canevas.fontTitre.i, strike=self.canevas.fontTitre.strike, outline=self.canevas.fontTitre.outline, shadow=self.canevas.fontTitre.shadow, condense=self.canevas.fontTitre.condense, color=self.canevas.fontTitre.color, extend=self.canevas.fontTitre.extend, sz=self.canevas.fontTitre.size, u=self.canevas.fontTitre.u, vertAlign=self.canevas.fontTitre.vertAlign, scheme=self.canevas.fontTitre.scheme)
                ws.cell(row=2,column = 4).value = self.canevas.titrePourcentage
                ws.cell(row =2, column = 4).alignment = Alignment(wrap_text=True)
                ws.cell(row=2,column = 4).fill = PatternFill(patternType=self.canevas.colors[1].patternType, fgColor=self.canevas.colors[1].fgColor, bgColor=self.canevas.colors[1].bgColor) 
                ws.cell(row=2,column=4).font=Font(name=self.canevas.fontPourcentage.name, charset=self.canevas.fontPourcentage.charset, family=self.canevas.fontPourcentage.family, b=self.canevas.fontPourcentage.b, i=self.canevas.fontPourcentage.i, strike=self.canevas.fontPourcentage.strike, outline=self.canevas.fontPourcentage.outline, shadow=self.canevas.fontPourcentage.shadow, condense=self.canevas.fontPourcentage.condense, color=self.canevas.fontPourcentage.color, extend=self.canevas.fontPourcentage.extend, sz=self.canevas.fontPourcentage.size, u=self.canevas.fontPourcentage.u, vertAlign=self.canevas.fontPourcentage.vertAlign, scheme=self.canevas.fontPourcentage.scheme)
                ws.cell(row=2,column = 4).fill = PatternFill(patternType=self.canevas.colors[1].patternType, fgColor=self.canevas.colors[1].fgColor, bgColor=self.canevas.colors[1].bgColor) 
                ws.cell(row=3,column = 6).value = self.canevas.titreNoteFinale
                ws.cell(row =3, column = 6).alignment = Alignment(wrap_text=True)
                ws.cell(row=3,column=6).font=Font(name=self.canevas.fontTitreNoteFinale.name, charset=self.canevas.fontTitreNoteFinale.charset, family=self.canevas.fontTitreNoteFinale.family, b=self.canevas.fontTitreNoteFinale.b, i=self.canevas.fontTitreNoteFinale.i, strike=self.canevas.fontTitreNoteFinale.strike, outline=self.canevas.fontTitreNoteFinale.outline, shadow=self.canevas.fontTitreNoteFinale.shadow, condense=self.canevas.fontTitreNoteFinale.condense, color=self.canevas.fontTitreNoteFinale.color, extend=self.canevas.fontTitreNoteFinale.extend, sz=self.canevas.fontTitreNoteFinale.size, u=self.canevas.fontTitreNoteFinale.u, vertAlign=self.canevas.fontTitreNoteFinale.vertAlign, scheme=self.canevas.fontTitreNoteFinale.scheme)
                i=4
                for appreciation in self.canevas.appreciations:
                    debutappreciation=i
                    #ws.cell(row = debutappreciation, column = 6).border = Border(self.canevas.border.left, self.canevas.border.right, self.canevas.border.top, self.canevas.border.bottom)
                    ws.cell(row =debutappreciation, column = 6).alignment = Alignment(wrap_text=True)
                    ws.cell(row=debutappreciation,column = 6).fill = PatternFill(patternType=self.canevas.colors[0].patternType, fgColor=self.canevas.colors[0].fgColor, bgColor=self.canevas.colors[0].bgColor)
                    ws.cell(row=debutappreciation,column=6).font=Font(name=self.canevas.fontNoteFinale.name, charset=self.canevas.fontNoteFinale.charset, family=self.canevas.fontNoteFinale.family, b=self.canevas.fontNoteFinale.b, i=self.canevas.fontNoteFinale.i, strike=self.canevas.fontNoteFinale.strike, outline=self.canevas.fontNoteFinale.outline, shadow=self.canevas.fontNoteFinale.shadow, condense=self.canevas.fontNoteFinale.condense, color=self.canevas.fontNoteFinale.color, extend=self.canevas.fontNoteFinale.extend, sz=self.canevas.fontNoteFinale.size, u=self.canevas.fontNoteFinale.u, vertAlign=self.canevas.fontNoteFinale.vertAlign, scheme=self.canevas.fontNoteFinale.scheme)
                    ws.merge_cells(start_row = i,start_column = 1, end_row = i, end_column = 2)
                    ws.cell(row=i,column = 1).value = appreciation.titre
                    ws.cell(row =i, column = 1).alignment = Alignment(wrap_text=True)
                    ws.cell(row=i,column=1).font=Font(name=self.canevas.fontAppreciation.name, charset=self.canevas.fontAppreciation.charset, family=self.canevas.fontAppreciation.family, b=self.canevas.fontAppreciation.b, i=self.canevas.fontAppreciation.i, strike=self.canevas.fontAppreciation.strike, outline=self.canevas.fontAppreciation.outline, shadow=self.canevas.fontAppreciation.shadow, condense=self.canevas.fontAppreciation.condense, color=self.canevas.fontAppreciation.color, extend=self.canevas.fontAppreciation.extend, sz=self.canevas.fontAppreciation.size, u=self.canevas.fontAppreciation.u, vertAlign=self.canevas.fontAppreciation.vertAlign, scheme=self.canevas.fontAppreciation.scheme)
                    ws.cell(row=i,column = 1).fill = PatternFill(patternType=self.canevas.colors[0].patternType, fgColor=self.canevas.colors[0].fgColor, bgColor=self.canevas.colors[0].bgColor)
                    ws.cell(row=i,column = 3).value = appreciation.nbPoints
                    ws.cell(row =i, column = 3).alignment = Alignment(wrap_text=True)
                    ws.cell(row=i,column=3).font=Font(name=self.canevas.fontNoteAppreciation.name, charset=self.canevas.fontNoteAppreciation.charset, family=self.canevas.fontNoteAppreciation.family, b=self.canevas.fontNoteAppreciation.b, i=self.canevas.fontNoteAppreciation.i, strike=self.canevas.fontNoteAppreciation.strike, outline=self.canevas.fontNoteAppreciation.outline, shadow=self.canevas.fontNoteAppreciation.shadow, condense=self.canevas.fontNoteAppreciation.condense, color=self.canevas.fontNoteAppreciation.color, extend=self.canevas.fontNoteAppreciation.extend, sz=self.canevas.fontNoteAppreciation.size, u=self.canevas.fontNoteAppreciation.u, vertAlign=self.canevas.fontNoteAppreciation.vertAlign, scheme=self.canevas.fontNoteAppreciation.scheme)
                    ws.cell(row=i,column = 3).fill = PatternFill(patternType=self.canevas.colors[0].patternType, fgColor=self.canevas.colors[0].fgColor, bgColor=self.canevas.colors[0].bgColor) 
                    ws.cell(row=i,column = 4).value = str(int(float(appreciation.partDeLaNote)*100)) + '%' 
                    ws.cell(row=i,column=4).font=Font(name=self.canevas.fontNotePourcentage.name, charset=self.canevas.fontNotePourcentage.charset, family=self.canevas.fontNotePourcentage.family, b=self.canevas.fontNotePourcentage.b, i=self.canevas.fontNotePourcentage.i, strike=self.canevas.fontNotePourcentage.strike, outline=self.canevas.fontNotePourcentage.outline, shadow=self.canevas.fontNotePourcentage.shadow, condense=self.canevas.fontNotePourcentage.condense, color=self.canevas.fontNotePourcentage.color, extend=self.canevas.fontNotePourcentage.extend, sz=self.canevas.fontNotePourcentage.size, u=self.canevas.fontNotePourcentage.u, vertAlign=self.canevas.fontNotePourcentage.vertAlign, scheme=self.canevas.fontNotePourcentage.scheme)
                    ws.cell(row=i,column = 4).fill = PatternFill(patternType=self.canevas.colors[1].patternType, fgColor=self.canevas.colors[1].fgColor, bgColor=self.canevas.colors[1].bgColor) 
                    i+=1
                    for critere in appreciation.criteres:
                        ws.merge_cells(start_row = i,start_column = 1, end_row =i, end_column = 2)
                        ws.cell(row=i,column = 1).value = critere.titre
                        ws.cell(row =i, column = 1).alignment = Alignment(wrap_text=True)
                        ws.cell(row=i,column=1).font=Font(name=self.canevas.fontTitreCritere.name, charset=self.canevas.fontTitreCritere.charset, family=self.canevas.fontTitreCritere.family, b=self.canevas.fontTitreCritere.b, i=self.canevas.fontTitreCritere.i, strike=self.canevas.fontTitreCritere.strike, outline=self.canevas.fontTitreCritere.outline, shadow=self.canevas.fontTitreCritere.shadow, condense=self.canevas.fontTitreCritere.condense, color=self.canevas.fontTitreCritere.color, extend=self.canevas.fontTitreCritere.extend, sz=self.canevas.fontTitreCritere.size, u=self.canevas.fontTitreCritere.u, vertAlign=self.canevas.fontTitreCritere.vertAlign, scheme=self.canevas.fontTitreCritere.scheme)
                        ws.cell(row=i,column = 3).value = critere.points
                        ws.cell(row =i, column = 3).alignment = Alignment(wrap_text=True)
                        ws.cell(row=i,column=3).font=Font(name=self.canevas.fontNoteCritere.name, charset=self.canevas.fontNoteCritere.charset, family=self.canevas.fontNoteCritere.family, b=self.canevas.fontNoteCritere.b, i=self.canevas.fontNoteCritere.i, strike=self.canevas.fontNoteCritere.strike, outline=self.canevas.fontNoteCritere.outline, shadow=self.canevas.fontNoteCritere.shadow, condense=self.canevas.fontNoteCritere.condense, color=self.canevas.fontNoteCritere.color, extend=self.canevas.fontNoteCritere.extend, sz=self.canevas.fontNoteCritere.size, u=self.canevas.fontNoteCritere.u, vertAlign=self.canevas.fontNoteCritere.vertAlign, scheme=self.canevas.fontNoteCritere.scheme)
                        if (critere.description == None):
                            i+=1
                        else:
                            j=8 #indice de colonnes parcourues à travers les jurés
                            for projet in self.projets:
                                ws.merge_cells(start_row = i,start_column = j, end_row = i+1, end_column = j)
                                j+=2
                            ws.merge_cells(start_row = i,start_column = 3, end_row = i+1, end_column = 3)
                            ws.row_dimensions[i+1].height = critere.hauteur
                            ws.cell(row=i+1,column = 2).value = critere.description
                            ws.cell(row =i+1, column = 2).alignment = Alignment(wrap_text=True)
                            ws.cell(row=i+1,column=2).font=Font(name=self.canevas.fontDescriptionCritere.name, charset=self.canevas.fontDescriptionCritere.charset, family=self.canevas.fontDescriptionCritere.family, b=self.canevas.fontDescriptionCritere.b, i=self.canevas.fontDescriptionCritere.i, strike=self.canevas.fontDescriptionCritere.strike, outline=self.canevas.fontDescriptionCritere.outline, shadow=self.canevas.fontDescriptionCritere.shadow, condense=self.canevas.fontDescriptionCritere.condense, color=self.canevas.fontDescriptionCritere.color, extend=self.canevas.fontDescriptionCritere.extend, sz=self.canevas.fontDescriptionCritere.size, u=self.canevas.fontDescriptionCritere.u, vertAlign=self.canevas.fontDescriptionCritere.vertAlign, scheme=self.canevas.fontDescriptionCritere.scheme)
                            i+=2
                    finappreciation=i-1
                    ws.merge_cells(start_row = debutappreciation+1,start_column = 6, end_row = finappreciation, end_column = 6)
                    j=8
                    debutprojet=j
                    for projet in self.projets:
                        ws.cell(row=debutappreciation,column=j).value='=SUM('+indextoletter(j)+str(debutappreciation+1)+':'+indextoletter(j)+str(finappreciation)+')'
                        ws.merge_cells(start_row = debutappreciation,start_column = j+1, end_row = finappreciation, end_column = j+1)
                        j+=2
                    finprojet=j
                    ws.cell(row=debutappreciation, column=6).value= "=AVERAGE("+indextoletter(debutprojet)+str(debutappreciation)+":"+indextoletter(finprojet-1)+str(debutappreciation)+")"
                    i+=1
                j=8
                numeroprojet=0
                for projet in self.projets:
                    ws.merge_cells(start_row = 2,start_column = j, end_row = 2, end_column = j+1)
                    ws.column_dimensions[indextoletter(j)].width = self.canevas.widthNoteJure
                    ws.column_dimensions[indextoletter(j+1)].width = self.canevas.widthCommentaireJure
                    ws.cell(row=2, column=j).value = projet.nom
                    ws.cell(row =2, column = j).alignment = Alignment(wrap_text=True)
                    ws.cell(row=2,column=j).font=Font(name=self.canevas.fontTitreJure.name, charset=self.canevas.fontTitreJure.charset, family=self.canevas.fontTitreJure.family, b=self.canevas.fontTitreJure.b, i=self.canevas.fontTitreJure.i, strike=self.canevas.fontTitreJure.strike, outline=self.canevas.fontTitreJure.outline, shadow=self.canevas.fontTitreJure.shadow, condense=self.canevas.fontTitreJure.condense, color=self.canevas.fontTitreJure.color, extend=self.canevas.fontTitreJure.extend, sz=self.canevas.fontTitreJure.size, u=self.canevas.fontTitreJure.u, vertAlign=self.canevas.fontTitreJure.vertAlign, scheme=self.canevas.fontTitreJure.scheme)
                    ws.cell(row=3, column=j).value = self.canevas.titreColonneNoteJure
                    ws.cell(row =3, column = j).alignment = Alignment(wrap_text=True)
                    ws.cell(row=3,column=j).font=Font(name=self.canevas.fontColonneNoteJure.name, charset=self.canevas.fontColonneNoteJure.charset, family=self.canevas.fontColonneNoteJure.family, b=self.canevas.fontColonneNoteJure.b, i=self.canevas.fontColonneNoteJure.i, strike=self.canevas.fontColonneNoteJure.strike, outline=self.canevas.fontColonneNoteJure.outline, shadow=self.canevas.fontColonneNoteJure.shadow, condense=self.canevas.fontColonneNoteJure.condense, color=self.canevas.fontColonneNoteJure.color, extend=self.canevas.fontColonneNoteJure.extend, sz=self.canevas.fontColonneNoteJure.size, u=self.canevas.fontColonneNoteJure.u, vertAlign=self.canevas.fontColonneNoteJure.vertAlign, scheme=self.canevas.fontColonneNoteJure.scheme)
                    ws.cell(row=3, column=j+1).value = self.canevas.titreColonneCommentaireJure
                    ws.cell(row =3, column = j+1).alignment = Alignment(wrap_text=True)
                    ws.cell(row=3,column=j+1).font=Font(name=self.canevas.fontColonneCommentaireJure.name, charset=self.canevas.fontColonneCommentaireJure.charset, family=self.canevas.fontColonneCommentaireJure.family, b=self.canevas.fontColonneCommentaireJure.b, i=self.canevas.fontColonneCommentaireJure.i, strike=self.canevas.fontColonneCommentaireJure.strike, outline=self.canevas.fontColonneCommentaireJure.outline, shadow=self.canevas.fontColonneCommentaireJure.shadow, condense=self.canevas.fontColonneCommentaireJure.condense, color=self.canevas.fontColonneCommentaireJure.color, extend=self.canevas.fontColonneCommentaireJure.extend, sz=self.canevas.fontColonneCommentaireJure.size, u=self.canevas.fontColonneCommentaireJure.u, vertAlign=self.canevas.fontColonneCommentaireJure.vertAlign, scheme=self.canevas.fontColonneCommentaireJure.scheme)
                    ws.cell(row=2,column = j).fill = PatternFill(patternType=self.canevas.colors[3+numeroprojet%2].patternType, fgColor=self.canevas.colors[3+numeroprojet%2].fgColor, bgColor=self.canevas.colors[3+numeroprojet%2].bgColor)
                    ws.cell(row=3,column = j).fill = PatternFill(patternType=self.canevas.colors[3+numeroprojet%2].patternType, fgColor=self.canevas.colors[3+numeroprojet%2].fgColor, bgColor=self.canevas.colors[3+numeroprojet%2].bgColor)
                    ws.cell(row=3,column = j+1).fill = PatternFill(patternType=self.canevas.colors[3+numeroprojet%2].patternType, fgColor=self.canevas.colors[3+numeroprojet%2].fgColor, bgColor=self.canevas.colors[3+numeroprojet%2].bgColor)
                    numeroprojet+=1
                    j+=2
        self.wbout.save("FichierSortie_parEtudiant.xlsx")
        
        
    def generate_sortie_etudiant(self):
        self.wbout = Workbook()
        self.generate_synthese_projet_etudiant()
        self.wbout.remove(self.wbout[self.wbout.sheetnames[0]])
        self.wbout.save("FichierSortie_parEtudiant.xlsx") 
        self.generate_projets_sheets_etudiant()
        self.generate_etudiant_sheets()
 
        
        
        